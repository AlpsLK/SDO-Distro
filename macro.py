# macro.py

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
import zipfile
import shutil
import os
import sys


FOLDER = None  # set via --input-dir or defaults to cwd


def sanitize_table_name(name: str) -> str:
    name = re.sub(r'\W+', '_', name)
    if not re.match(r'[A-Za-z_]', name):
        name = "_" + name
    return name[:255]


def colletters(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def find_used_range(ws):
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    last_row = 0
    for r in range(1, max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, max_col + 1)):
            last_row = r
    if last_row == 0:
        return None

    last_col = 0
    for c in range(1, max_col + 1):
        if any(ws.cell(r, c).value not in (None, "") for r in range(1, last_row + 1)):
            last_col = c

    first_row = 1
    for r in range(1, last_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, last_col + 1)):
            first_row = r
            break

    first_col = 1
    for c in range(1, last_col + 1):
        if any(ws.cell(r, c).value not in (None, "") for r in range(first_row, last_row + 1)):
            first_col = c
            break

    return (first_row, first_col, last_row, last_col)


def add_table_if_missing(ws):
    if ws._tables:
        return False

    used = find_used_range(ws)
    if not used:
        return False

    r1, c1, r2, c2 = used
    ref = f"{colletters(c1)}{r1}:{colletters(c2)}{r2}"

    base = sanitize_table_name(f"{ws.title}_Table")
    existing = {t.displayName for t in ws._tables}
    name = base
    i = 2
    while name in existing:
        name = sanitize_table_name(f"{base}_{i}")
        i += 1

    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tbl)
    return True


def process_excel_files(folder: Path):
    files = [
        p for p in folder.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    ]

    if not files:
        print("No .xlsx files found to process.")
        return []

    processed = []

    for f in files:
        print(f"Processing: {f.name}")
        try:
            wb = load_workbook(f)
        except Exception as e:
            print(f"SKIP (open failed): {f.name} -> {e}")
            continue

        changed = False
        for ws in wb.worksheets:
            if add_table_if_missing(ws):
                changed = True
                print(f"  + Added table on sheet '{ws.title}'")

        try:
            wb.save(f)
            print(f"  SAVED -> {f.name}")
            processed.append(f)
        except Exception as e:
            print(f"ERROR saving {f.name}: {e}")

    return processed


def detect_service_name(current_folder: Path) -> str:
    name = current_folder.name
    if name.endswith("-results"):
        return name[:-8]
    return name


def finalize_outputs(service: str, service_results_dir: Path, excel_files: list):
    """
    Move xlsx files and create zip archive.
    JSON results are already written to Zips_Archive/json/YYYY-MM/<service>-results.
    This function creates:
      Zips_Archive/excell/YYYY-MM/<service>/*.xlsx
      Zips_Archive/zip/YYYY-MM/<service>.zip
    """
    # Zips_Archive
    try:
        base_dir = service_results_dir.parents[2]
    except IndexError:
        base_dir = service_results_dir.parent

    month_segment = service_results_dir.parent.name

    json_dir = base_dir / "json" / month_segment
    excell_dir = base_dir / "excell" / month_segment
    zip_dir = base_dir / "zip" / month_segment

    json_dir.mkdir(parents=True, exist_ok=True)
    excell_dir.mkdir(parents=True, exist_ok=True)
    zip_dir.mkdir(parents=True, exist_ok=True)

    excel_service_folder = excell_dir / service
    excel_service_folder.mkdir(exist_ok=True)

    for f in excel_files:
        shutil.move(str(f), excel_service_folder / f.name)

    zip_path = zip_dir / f"{service}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(excel_service_folder):
            for file in files:
                full = Path(root) / file
                rel = full.relative_to(excel_service_folder)
                zf.write(full, arcname=rel)

    print(f"ZIP CREATED -> {zip_path.name}")
    print(f"JSON RESULTS at {service_results_dir}")


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", dest="input_dir", default=None,
                        help="Service results directory to process")
    args, _ = parser.parse_known_args()

    if args.input_dir:
        current = Path(args.input_dir).resolve()
    elif FOLDER is not None:
        current = FOLDER.resolve()
    else:
        current = Path(".").resolve()

    service = detect_service_name(current)
    print(f"Detected service: {service}")

    excel_files = process_excel_files(current)
    finalize_outputs(service, current, excel_files)

    print("\nDONE.")


if __name__ == "__main__":
    main()
 
